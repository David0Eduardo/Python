import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from ofxparse import OfxParser
import tempfile
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Definições dos caminhos dos dicionários diretamente no código
CAMINHO_DICIONARIO_CREDIT = r'Y:\Informações Caixas\Medicina e Odontologia\REPOSITORIO\DICIONARIOS\credit'
CAMINHO_DICIONARIO_DEBIT = r'Y:\Informações Caixas\Medicina e Odontologia\REPOSITORIO\DICIONARIOS\debit'

# Inicializa os dicionários de frases
frases_por_dicionario_credit = {}
frases_por_dicionario_debit = {}

# Função para atualizar o histórico no Listbox
def atualizar_historico(mensagem):
    if 'listbox_files' in globals():
        listbox_files.insert(tk.END, mensagem)
        listbox_files.yview(tk.END)  # Rola para a última linha
    else:
        print(mensagem)  # Caso o Listbox não esteja disponível, imprime a mensagem no console

def carregar_dicionarios():
    global frases_por_dicionario_credit, frases_por_dicionario_debit
    
    # Carregar dicionário de crédito
    for arquivo in os.listdir(CAMINHO_DICIONARIO_CREDIT):
        if arquivo.endswith(".txt"):
            caminho_arquivo_txt = os.path.join(CAMINHO_DICIONARIO_CREDIT, arquivo)
            categoria = os.path.splitext(arquivo)[0]  # Nome do arquivo sem extensão
            
            try:
                with open(caminho_arquivo_txt, 'r', encoding='utf-8') as file:
                    frases = [linha.strip() for linha in file if linha.strip()]
                    frases_por_dicionario_credit[categoria] = frases
                    atualizar_historico(f"Frases carregadas do dicionário '{categoria}' para tipo 'credit'")
            except Exception as e:
                atualizar_historico(f"Erro ao ler o arquivo de frases '{caminho_arquivo_txt}': {e}")

    # Carregar dicionário de débito
    for arquivo in os.listdir(CAMINHO_DICIONARIO_DEBIT):
        if arquivo.endswith(".txt"):
            caminho_arquivo_txt = os.path.join(CAMINHO_DICIONARIO_DEBIT, arquivo)
            categoria = os.path.splitext(arquivo)[0]  # Nome do arquivo sem extensão
            
            try:
                with open(caminho_arquivo_txt, 'r', encoding='utf-8') as file:
                    frases = [linha.strip() for linha in file if linha.strip()]
                    frases_por_dicionario_debit[categoria] = frases
                    atualizar_historico(f"Frases carregadas do dicionário '{categoria}' para tipo 'debit'")
            except Exception as e:
                atualizar_historico(f"Erro ao ler o arquivo de frases '{caminho_arquivo_txt}': {e}")

# Função para salvar os dados classificados em um arquivo Excel
def salvar_dados_classificados(dados_classificados, caminho_saida):
    try:
        if not dados_classificados:
            atualizar_historico("Nenhum dado para salvar.")
            messagebox.showinfo("Aviso", "Nenhum dado para salvar.")
            return
        
        # Definindo a nova ordem das colunas com "Classificação" como a última
        df = pd.DataFrame(dados_classificados, columns=[
            "Data", "Valor", "Banco", "Agencia", "Conta", "Nome do Arquivo", "Tipo", "Memo", "ID", "Classificação"
        ])
        
        # Adicionar novas colunas com fórmulas de ajuste
        df["Data Ajustada"] = None
        df["Valor Ajustado"] = None
        df["Valor Absoluto"] = None
        
        # Reordenar as colunas para que a coluna "Classificação" fique por último
        colunas = [
            "Data", "Valor", "Banco", "Agencia", "Conta", "Nome do Arquivo", "Tipo", "Memo", "ID", "Data Ajustada", "Valor Ajustado", "Valor Absoluto", "Classificação"
        ]
        df = df[colunas]
        
        # Salvar DataFrame como arquivo Excel
        df.to_excel(caminho_saida, index=False)
        atualizar_historico(f"Arquivo Excel salvo em '{caminho_saida}'")
        
        # Manipular o arquivo Excel para adicionar fórmulas e ocultar colunas
        wb = load_workbook(caminho_saida)
        ws = wb.active
        
        # Adicionar fórmulas para cada linha
        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para ignorar o cabeçalho
            ws[f"J{row}"] = f'=DATEVALUE(A{row})'  # Data Ajustada
            ws[f"K{row}"] = f'=VALUE(SUBSTITUTE(B{row},".",","))'  # Valor Ajustado
            ws[f"L{row}"] = f'=ABS(K{row})'  # Valor Absoluto
        
        # Ocultar as colunas especificadas
        for col_letter in ['A', 'B', 'I', 'N', 'O']:
            ws.column_dimensions[col_letter].hidden = True

        wb.save(caminho_saida)
        atualizar_historico(f"Fórmulas adicionadas e colunas ocultas no arquivo Excel '{caminho_saida}'")
        messagebox.showinfo("Sucesso", f"Arquivo processado e salvo como '{caminho_saida}'")
    
    except Exception as e:
        atualizar_historico(f"Erro ao salvar o arquivo Excel: {e}")
        messagebox.showerror("Erro", f"Ocorreu um erro ao salvar o arquivo Excel: {e}")

def processar_dados_ofx(folder_paths):
    dados_classificados = []

    if not frases_por_dicionario_credit and not frases_por_dicionario_debit:
        atualizar_historico("Nenhuma frase de dicionário carregada.")
        return dados_classificados

    for folder_path in folder_paths:
        for filename in os.listdir(folder_path):
            if filename.endswith(".ofx"):
                ofx_file_path = os.path.join(folder_path, filename)
                try:
                    with open(ofx_file_path, 'rb') as file:
                        ofx_data = file.read().decode('utf-8', errors='ignore')
                except Exception as e:
                    atualizar_historico(f"Erro ao ler o arquivo OFX '{filename}': {e}")
                    continue
                
                try:
                    with tempfile.NamedTemporaryFile(delete=False) as temp_file:
                        temp_file.write(ofx_data.encode('utf-8'))
                        temp_file_path = temp_file.name
                    ofx = OfxParser.parse(open(temp_file_path, 'rb'))
                    os.unlink(temp_file_path)
                except Exception as e:
                    atualizar_historico(f"Erro ao processar o arquivo OFX '{filename}': {e}")
                    continue

                atualizar_historico(f"Arquivo OFX processado: {filename}")
                
                if not hasattr(ofx.account, 'statement') or not hasattr(ofx.account.statement, 'transactions'):
                    atualizar_historico(f"Transações não encontradas no OFX '{filename}'")
                    continue

                for transaction in ofx.account.statement.transactions:
                    memo = transaction.memo if transaction.memo else ''  # Inicialize memo
                    valor = transaction.amount  # Valor da transação

                    # Determina o tipo da transação com base no valor
                    if valor > 0:
                        tipo_transacao = 'CREDIT'
                        dicionario = frases_por_dicionario_credit
                        classificacao = 'Receita de Vendas'
                    elif valor < 0:
                        tipo_transacao = 'DEBIT'
                        dicionario = frases_por_dicionario_debit
                        classificacao = ''  # Para 'DEBIT', não é necessário definir uma classificação padrão.
                    else:
                        atualizar_historico(f"Valor da transação é zero, não classificado: {transaction}")
                        continue

                    if dicionario:
                        for categoria, frases in dicionario.items():
                            if any(frase in str(memo) for frase in frases):
                                classificacao = categoria
                                break

                    dados_classificados.append([
                        transaction.date.strftime("%Y-%m-%d"),
                        valor,
                        getattr(ofx.account, 'routing_number', 'N/A'),
                        getattr(ofx.account, 'branch_id', 'N/A'),
                        getattr(ofx.account, 'account_id', 'N/A'),
                        filename,
                        tipo_transacao,
                        memo,
                        transaction.id if hasattr(transaction, 'id') else 'N/A',
                        classificacao
                    ])

    atualizar_historico(f"Dados classificados: {len(dados_classificados)} transações")
    
    return dados_classificados

# Função para selecionar pastas OFX
def selecionar_pastas_ofx():
    global folder_paths
    folder_paths = []  # Lista para armazenar múltiplos caminhos de pastas
    
    root = tk.Tk()
    root.withdraw()
    
    while True:
        folder_path = filedialog.askdirectory(title="Selecione uma pasta contendo arquivos OFX")
        if folder_path:
            folder_paths.append(folder_path)  # Adiciona o caminho da pasta à lista
            atualizar_historico(f"Pasta OFX selecionada: {folder_path}")
        else:
            break

# Função para selecionar o local para salvar o arquivo Excel e iniciar o processamento
def selecionar_pasta_salvar_excel():
    caminho_saida = filedialog.asksaveasfilename(defaultextension=".xlsx")
    if caminho_saida:
        dados_classificados = processar_dados_ofx(folder_paths)
        if dados_classificados:
            salvar_dados_classificados(dados_classificados, caminho_saida)

# Interface Tkinter
root = tk.Tk()
root.title("Processamento e Classificação de OFX")

# Frame para o conteúdo principal
frame_main = tk.Frame(root)
frame_main.pack(fill=tk.BOTH, expand=True)

# Frame para os botões
frame_buttons = tk.Frame(frame_main)
frame_buttons.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)

# Botão para iniciar a seleção das pastas OFX
button_selecionar_pastas = tk.Button(frame_buttons, text="Selecionar Pastas OFX", command=selecionar_pastas_ofx, width=20, height=5)
button_selecionar_pastas.pack(pady=10, padx=10)

# Botão para selecionar onde salvar o arquivo Excel
button_salvar_excel = tk.Button(frame_buttons, text="Salvar Excel e Processar", command=selecionar_pasta_salvar_excel, width=20, height=5)
button_salvar_excel.pack(pady=10, padx=10)

# Frame para a lista de histórico
frame_listbox = tk.Frame(frame_main)
frame_listbox.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=10, pady=10)

# Criação do Listbox e scrollbar
listbox_files = tk.Listbox(frame_listbox, selectmode=tk.SINGLE, width=50, height=20)
scrollbar = tk.Scrollbar(frame_listbox, orient=tk.VERTICAL, command=listbox_files.yview)
listbox_files.config(yscrollcommand=scrollbar.set)

listbox_files.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

root.geometry("800x600")

# Carregar dicionários após inicialização da interface gráfica
carregar_dicionarios()

root.mainloop()
